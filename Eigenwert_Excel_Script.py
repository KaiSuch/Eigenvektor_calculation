import xlrd
import numpy as np
import pandas as pd
import openpyxl as op


# ------------------------------------------------- Eingabebereich -----------------------------------------------------

# enter the names of your excel sheets
NamedesDokumentes = "Kopie von Wissensdatenbank - Fuzzy_13.xlsx"
NamedesTabellenblattes = "Paarvergleich"
NamedesneuenDokuments = "Wissensdatenbank - Fuzzy - ausgefüllt.xlsx"

# ------------------------------------------------- Matrixauswahl -----------------------------------------------------

# each block represents a matrix from the top left to the bottom right corner, choose how much you need and delete the others

X1 = "F"
Y1 = 4
X2 = "G"
Y2 = 5

X22 = "F"
Y22 = 9
X33 = "H"
Y33 = 11

X3 = "F"
Y3 = 15
X4 = "K"
Y4 = 20

X5 = "F"
Y5 = 24
X6 = "L"
Y6 = 30

X7 = "F"
Y7 = 34
X8 = "M"
Y8 = 41

X9 = "F"
Y9 = 44
X10 = "O"
Y10 = 53

X11 = "F"
Y11 = 56
X12 = "K"
Y12 = 61

X13 = "F"
Y13 = 65
X14 = "L"
Y14 = 71

X15 = "F"
Y15 = 75
X16 = "H"
Y16 = 77

# ------------------------------------------------- Setup, ab hier keine Änderungen vornehmen! -----------------------------------------------------

workbook = xlrd.open_workbook((NamedesDokumentes))
worksheet = workbook.sheet_by_name(NamedesTabellenblattes)
wb = op.load_workbook(NamedesDokumentes)
sheet = wb[NamedesTabellenblattes]


# ------------------------------------------------- Funktionen -----------------------------------------------------

# Wandelt die Spaltennummern in Zahlen um
def col_to_num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num

# Trägt den normierten Eigenvektor an die entsprechende Stelle ein
def NormierteEigenvektor_eintragen(Array, StartY, StartX):
    NormierterEigenvektor = Array.real
    t = 0
    for _ in NormierterEigenvektor:
        sheet.cell(row=StartY +1, column=StartX + 2).value = NormierterEigenvektor[t]
        StartY = StartY + 1
        t = t + 1

# Trägt die absolut Werte des Eigenvektors an die enstpechende Stelle ein
def AbsoluteEigenwerte_eintragen(Array, StarY, StartX):
        AbsoluteEigenwerte = Array.real
        t = 0
        for _ in AbsoluteEigenwerte:
            sheet.cell(row=StarY + 1, column=StartX + 1).value = AbsoluteEigenwerte[t]
            StarY = StarY + 1
            t = t + 1

# Trägt den größten Eigenwert an die enstpechende Stelle ein
def Eigenwert_eintragen(GrößterEigenwert, X2, Y2):
    sheet.cell(row=Y2 + 1, column=X2 + 2,).value = GrößterEigenwert

# Trägt den Inkonsistenzwert an die enstpechende Stelle ein
def Inkonsistenzwert_eintragen(inkonsistenzwert, X2, Y1):
    sheet.cell(row=Y1+1, column=X2 +3).value = inkonsistenzwert

# Hauptfunktion die alle Berechungen ausführt
def Matrixaufspannen(X1, X2, Y1, Y2):
    Y1 = Y1 - 1
    Y2 = Y2
    X1 = col_to_num(X1) - 1
    X2 = col_to_num(X2)
    table = list()
    record = list()
    for x in range(Y1, Y2):
        for y in range(X1, X2):
            record.append(worksheet.cell(x, y).value)
        table.append(record)
        record = []
        x += 1
    Eigenwerte = np.linalg.eigvals(table)   # Ermittelt die Eigenvektoren
    GrößterEigenwert = max(Eigenwerte)  # Wird nicht nach Absolutwerten Berechnet!
    IndexEigenwert = pd.Series(Eigenwerte).idxmax()     #Ermittelt den Index des höchsten Eigenwertes
    Inkonsistenzwert = (GrößterEigenwert - len(Eigenwerte))/(len(Eigenwerte) - 1)    # Berechnung des Inkonsistenzwertes
    Eigenvektoren = np.linalg.eig(table)[1]  # Eigenwerte und Eigenvektoren in einer Ausgabe
    Eigenvektor = Eigenvektoren[:, IndexEigenwert]
    AbsolteWerteEigenvektor = np.absolute(Eigenvektor)  # Absolue Eigenwerte berechnen, Betrag jedes Eintrags
    SummeEigenverktor = np.sum(AbsolteWerteEigenvektor)  # Summe der Eigenwerte
    NormierterEigenvektor = AbsolteWerteEigenvektor / SummeEigenverktor  # Normieren des Eigenvektors
    print("")
    print("-----------------------------------------")
    print(" Matrix: : \n", table)
    print("-----------------------------------------")
    print("")
    print("-----------------------------------------")
    print("Eigenwerte : \n", Eigenwerte)
    print("Alle Eigenvektoren : \n", Eigenvektoren)
    print("-----------------------------------------")
    print("")
    print("-----------------------------------------")
    print("Eigenvektor (Betragsmäßig) : \t", AbsolteWerteEigenvektor)
    print("Normierter Eigenvektor :\t \t", NormierterEigenvektor)
    print("-----------------------------------------")
    print("")
    print("-----------------------------------------")
    print("Index höchster Eigenwert :\t", pd.Series(Eigenwerte).idxmax())
    print("Größter Eigenwert :\t \t \t", GrößterEigenwert)
    print("Inkonsistenzwert :\t \t \t", Inkonsistenzwert)
    print("-----------------------------------------")
    print("___________________________________________________________________________________________________________________________________________________________________________________________________________________________________")
    print("___________________________________________________________________________________________________________________________________________________________________________________________________________________________________")
    GrößterEigenwert = GrößterEigenwert.real
    Eigenwert_eintragen(GrößterEigenwert, X2, Y2)
    AbsolteWerteEigenvektor = AbsolteWerteEigenvektor.real
    AbsoluteEigenwerte_eintragen(AbsolteWerteEigenvektor, Y1, X2)
    NormierterEigenvektor = NormierterEigenvektor.real
    NormierteEigenvektor_eintragen(NormierterEigenvektor, Y1, X2)
    Inkonsistenzwert = Inkonsistenzwert.real     #Inkonsistenzwert eintragen
    Inkonsistenzwert_eintragen(Inkonsistenzwert, X2, Y1)

# ------------------------------------------------- Funktionen starten -----------------------------------------------------

Matrixaufspannen(X1, X2, Y1, Y2)
Matrixaufspannen(X3, X4, Y3, Y4)
Matrixaufspannen(X5, X6, Y5, Y6)
Matrixaufspannen(X7, X8, Y7, Y8)
Matrixaufspannen(X9, X10, Y9, Y10)
Matrixaufspannen(X11, X12, Y11, Y12)
Matrixaufspannen(X13, X14, Y13, Y14)
Matrixaufspannen(X15, X16, Y15, Y16)

# ------------------------------------------------- Speichern -----------------------------------------------------
wb.save(NamedesneuenDokuments)
